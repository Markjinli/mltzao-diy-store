import json
import os
import requests
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMG_DIR = os.path.join(BASE_DIR, "images")
JSON_PATH = os.path.join(BASE_DIR, "products.json")
EXCEL_PATH = os.path.join(BASE_DIR, "mltzao_diy_store_configs.xlsx")

os.makedirs(IMG_DIR, exist_ok=True)

# Load data
with open(JSON_PATH, "r", encoding="utf-8") as f:
    products = json.load(f)

print(f"Loaded {len(products)} products")

# Download images
def download_image(url, diyId):
    if not url:
        return ""
    ext = ".jpg"
    if ".png" in url:
        ext = ".png"
    filename = f"{diyId}{ext}"
    filepath = os.path.join(IMG_DIR, filename)
    if os.path.exists(filepath):
        return filename
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.mltzao.com/"
        }
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        with open(filepath, "wb") as f:
            f.write(resp.content)
        print(f"  Downloaded: {filename} ({len(resp.content)} bytes)")
        return filename
    except Exception as e:
        print(f"  Failed to download {diyId}: {e}")
        return ""

print("\nDownloading images...")
img_map = {}
for p in products:
    fname = download_image(p["imgUrl"], p["diyId"])
    img_map[p["diyId"]] = fname

# Create Excel
print("\nGenerating Excel...")
wb = Workbook()
ws = wb.active
ws.title = "主机配置列表"

# Headers
headers = [
    "序号", "配置名称", "编号", "价格(元)", "鲁大师跑分",
    "机箱", "CPU", "显卡", "主板", "内存", "硬盘", "散热器", "电源",
    "预计发货日期", "图片文件"
]

# Style
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for idx, p in enumerate(products, 1):
    specs = p.get("specs", {})
    row_data = [
        idx,
        p["name"],
        p["diyId"],
        int(p["price"]) if p["price"] else 0,
        int(p["score"]) if p["score"] else 0,
        specs.get("机箱", ""),
        specs.get("CPU", ""),
        specs.get("显卡", ""),
        specs.get("主板", ""),
        specs.get("内存", ""),
        specs.get("硬盘", ""),
        specs.get("散热器", ""),
        specs.get("电源", ""),
        p["deliveryDate"],
        img_map.get(p["diyId"], "")
    ]
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=idx + 1, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        # Price column formatted
        if col == 4 and isinstance(value, int) and value > 0:
            cell.number_format = '#,##0'
        if col == 5 and isinstance(value, int) and value > 0:
            cell.number_format = '#,##0'

# Set column widths
col_widths = [6, 35, 14, 12, 12, 30, 25, 30, 25, 30, 25, 30, 35, 14, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze first row
ws.freeze_panes = "A2"

# Auto filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(products) + 1}"

wb.save(EXCEL_PATH)
print(f"\nExcel saved to: {EXCEL_PATH}")
print(f"Images saved to: {IMG_DIR}")
print(f"Total products: {len(products)}")
