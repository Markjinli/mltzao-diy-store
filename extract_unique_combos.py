import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(BASE_DIR, "products_full.json")
EXCEL_PATH = os.path.join(BASE_DIR, "mltzao_unique_combos.xlsx")

with open(JSON_PATH, "r", encoding="utf-8") as f:
    products = json.load(f)

print(f"Loaded {len(products)} products")


def extract_gpu_core(gpu_name):
    """从显卡名称中提取GPU核心型号，如 RTX 5060, RTX 5060 Ti, RTX 3050 等"""
    if not gpu_name:
        return "未知"

    # Pattern: RTX 5090 D, RTX 5090, RTX 5080, RTX 5070 Ti, RTX 5070, RTX 5060 Ti, RTX 5060, RTX 5050
    # RTX 4090, RTX 4080S, RTX 4070 Ti S, RTX 4070S, RTX 4070, RTX 4060 Ti, RTX 4060
    # RTX 3060, RTX 3050
    # RX 9070, RX 9070 XT, RX 7600, RX 7600 XT, RX 6750 GRE

    patterns = [
        # NVIDIA RTX 50 series
        (r'RTX\s*5090\s*D', 'RTX 5090 D'),
        (r'RTX\s*5090', 'RTX 5090'),
        (r'RTX\s*5080', 'RTX 5080'),
        (r'RTX\s*5070\s*Ti', 'RTX 5070 Ti'),
        (r'RTX\s*5070', 'RTX 5070'),
        (r'RTX\s*5060\s*Ti', 'RTX 5060 Ti'),
        (r'RTX\s*5060', 'RTX 5060'),
        (r'RTX\s*5050', 'RTX 5050'),
        # NVIDIA RTX 40 series
        (r'RTX\s*4090', 'RTX 4090'),
        (r'RTX\s*4080\s*S', 'RTX 4080S'),
        (r'RTX\s*4080', 'RTX 4080'),
        (r'RTX\s*4070\s*Ti\s*S', 'RTX 4070 Ti S'),
        (r'RTX\s*4070\s*Ti', 'RTX 4070 Ti'),
        (r'RTX\s*4070\s*S', 'RTX 4070S'),
        (r'RTX\s*4070', 'RTX 4070'),
        (r'RTX\s*4060\s*Ti', 'RTX 4060 Ti'),
        (r'RTX\s*4060', 'RTX 4060'),
        # NVIDIA RTX 30 series
        (r'RTX\s*3060', 'RTX 3060'),
        (r'RTX\s*3050', 'RTX 3050'),
        # AMD RX 9000 series
        (r'RX\s*9070\s*XT', 'RX 9070 XT'),
        (r'RX\s*9070', 'RX 9070'),
        (r'RX\s*9060\s*XT', 'RX 9060 XT'),
        (r'RX\s*9060', 'RX 9060'),
        # AMD RX 7000 series
        (r'RX\s*7900\s*XTX', 'RX 7900 XTX'),
        (r'RX\s*7900\s*XT', 'RX 7900 XT'),
        (r'RX\s*7800\s*XT', 'RX 7800 XT'),
        (r'RX\s*7700\s*XT', 'RX 7700 XT'),
        (r'RX\s*7600\s*XT', 'RX 7600 XT'),
        (r'RX\s*7600', 'RX 7600'),
        # AMD RX 6000 series
        (r'RX\s*6950\s*XT', 'RX 6950 XT'),
        (r'RX\s*7650\s*GRE', 'RX 7650 GRE'),
        (r'RX\s*6750\s*GRE', 'RX 6750 GRE'),
        (r'RX\s*6650\s*XT', 'RX 6650 XT'),
        (r'RX\s*6600', 'RX 6600'),
    ]

    for pattern, core_name in patterns:
        if re.search(pattern, gpu_name, re.IGNORECASE):
            return core_name

    # Fallback: try to extract any RTX/RX pattern
    m = re.search(r'(RTX\s*\d{4}\s*(?:Ti|S|D)?)', gpu_name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    m = re.search(r'(RX\s*\d{4}\s*(?:XT|GRE)?)', gpu_name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    return gpu_name


def extract_ram_spec(mem_name):
    """从内存名称中提取容量+频率，如 '16G DDR4 3200', '32G DDR5 6000'"""
    if not mem_name:
        return "未知"

    # Extract DDR type: DDR4, DDR5
    ddr_match = re.search(r'(DDR[45])', mem_name, re.IGNORECASE)
    ddr_type = ddr_match.group(1).upper() if ddr_match else ""

    # Handle dual channel: "24Gx2", "16G*2", "16GB*2" → total capacity
    dual_match = re.search(r'(\d+)\s*[Gg](?:B)?\s*[x\*]\s*(\d+)', mem_name)
    if dual_match:
        single = int(dual_match.group(1))
        count = int(dual_match.group(2))
        total = single * count
        capacity = f"{total}G"
    else:
        # Extract capacity: 8G, 16G, 24G, 32G, 48G, 64G, etc.
        cap_match = re.search(r'(\d+)\s*[Gg](?:B)?', mem_name)
        capacity = cap_match.group(1) + "G" if cap_match else ""

    # Extract frequency: 3200, 3600, 5600, 6000, 6400, 7200, etc.
    # Look for 4-digit numbers that are memory frequencies
    freq_match = re.search(r'(\d{4})\s*(?:MHz)?', mem_name)
    freq = freq_match.group(1) if freq_match else ""

    if capacity and ddr_type and freq:
        return f"{capacity} {ddr_type} {freq}"
    elif capacity and ddr_type:
        return f"{capacity} {ddr_type}"
    elif capacity and freq:
        return f"{capacity} DDR5 {freq}"  # Default to DDR5 if no DDR type but has freq
    elif capacity:
        return capacity
    return mem_name


def is_diy(computer_no):
    """判断编号是否为DIY类型"""
    return computer_no and computer_no.startswith("diy")


# Process all products
results = []
for p in products:
    specs = p.get("specs", {})
    cpu = specs.get("CPU", "")
    gpu = specs.get("显卡", "")
    mem = specs.get("内存", "")

    if not cpu or not gpu or not mem:
        continue

    gpu_core = extract_gpu_core(gpu)
    ram_spec = extract_ram_spec(mem)

    key = f"{cpu}|{gpu_core}|{ram_spec}"

    results.append({
        "key": key,
        "cpu": cpu,
        "gpu_core": gpu_core,
        "gpu_full": gpu,
        "ram_spec": ram_spec,
        "ram_full": mem,
        "product": p
    })

print(f"Total valid configs: {len(results)}")

# Group by key, apply selection rules
from collections import defaultdict
groups = defaultdict(list)
for r in results:
    groups[r["key"]].append(r)

final_products = []
for key, items in groups.items():
    # Sort: DIY first, then by price descending
    items.sort(key=lambda x: (
        1 if is_diy(x["product"].get("computerNo", "")) else 0,
        x["product"].get("price", 0)
    ), reverse=True)
    final_products.append(items[0])

# Sort final by price
final_products.sort(key=lambda x: x["product"].get("price", 0))

print(f"Unique combos: {len(final_products)}")

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "去重配置组合"

headers = [
    "序号", "配置名称", "编号", "价格(元)", "鲁大师跑分",
    "CPU", "GPU核心", "显卡(完整)", "内存规格", "内存(完整)",
    "机箱", "主板", "硬盘", "CPU散热", "电源",
    "发货时效"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for idx, item in enumerate(final_products, 1):
    p = item["product"]
    specs = p.get("specs", {})
    row_data = [
        idx,
        p.get("name", ""),
        p.get("computerNo", ""),
        p.get("price", 0),
        p.get("score", 0),
        item["cpu"],
        item["gpu_core"],
        item["gpu_full"],
        item["ram_spec"],
        item["ram_full"],
        specs.get("机箱", ""),
        specs.get("主板", ""),
        specs.get("硬盘", ""),
        specs.get("CPU散热", specs.get("散热器", "")),
        specs.get("电源", ""),
        p.get("deliveryDate", ""),
    ]
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=idx + 1, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col == 4 and isinstance(value, (int, float)) and value > 0:
            cell.number_format = '#,##0'
        if col == 5 and isinstance(value, (int, float)) and value > 0:
            cell.number_format = '#,##0'

col_widths = [6, 40, 14, 12, 12, 28, 16, 35, 18, 35, 35, 28, 28, 30, 40, 14]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(final_products) + 1}"

wb.save(EXCEL_PATH)
print(f"\nExcel saved: {EXCEL_PATH}")

# Print summary stats
cpu_set = set(item["cpu"] for item in final_products)
gpu_set = set(item["gpu_core"] for item in final_products)
ram_set = set(item["ram_spec"] for item in final_products)
print(f"\n--- Summary ---")
print(f"Unique CPUs: {len(cpu_set)}")
print(f"Unique GPU cores: {len(gpu_set)}")
print(f"Unique RAM specs: {len(ram_set)}")
print(f"Total unique combos: {len(final_products)}")
